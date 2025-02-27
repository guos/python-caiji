from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time
import _thread

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	time.sleep(4)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, cas, products):
	print(str(len(products))+ cas + url)
	
	productListHtml = getRenderdHtmlFromUrl(url)
	tempPinfo = {
		"cas":cas
	}
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	tempPinfo["Product Name"] = getNodeText(sope.find("h1", attrs={"class":"m-zero p-zero"}))
	ths = sope.find_all("th")
	for th in ths:
		title = getNodeText(th)
		if title == "Molecular Formula":
			tempPinfo["Molecular Formula"] = getNodeText(th.findNextSibling("td"))
		if title == "Synonyms":
			tempPinfo["Synonyms"] = getNodeText(th.findNextSibling("td"))
		if title == "Molecular Weight":
			tempPinfo["Molecular Weight"] = getNodeText(th.findNextSibling("td"))

	
	IUPACName = sope.find("section",attrs={"id":"IUPAC-Name"})
	InChI = sope.find("section",attrs={"id":"InChI"})
	InChIKey = sope.find("section",attrs={"id":"InChI-Key"})
	CanonicalSMILES = sope.find("section",attrs={"id":"Canonical-SMILES"})
	CAS = sope.find("section",attrs={"id":"CAS"})
	MolecularFormula = sope.find("section",attrs={"id":"Molecular-Formula"})
	if IUPACName != None:
		tempPinfo["IUPACName"] = getNodeText(IUPACName.find("div", attrs={"class":"section-content"}))
	if InChI != None:
		tempPinfo["InChI"] = getNodeText(InChI.find("div", attrs={"class":"section-content"}))
	if InChIKey != None:
		tempPinfo["InChIKey"] = getNodeText(InChIKey.find("div", attrs={"class":"section-content"}))
	if CanonicalSMILES != None:
		tempPinfo["CanonicalSMILES"] = getNodeText(CanonicalSMILES.find("div", attrs={"class":"section-content"}))
	if CAS != None:
		tempPinfo["CAS1"] = getNodeText(CAS.find("div", attrs={"class":"section-content"}))
	if MolecularFormula != None:
		tempPinfo["MolecularFormula"] = getNodeText(MolecularFormula.find("div", attrs={"class":"section-content"}))
	# print(tempPinfo["IUPACName"])
	products.append(tempPinfo.copy())

def getProductList(cas, products):
	productListHtml = getHtmlFromUrl("https://www.ncbi.nlm.nih.gov/pccompound/?term="+cas)
	if productListHtml!=None:
		sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
		pros = sope.find_all("div", attrs={"class":"rprt"})
		if len(pros) > 0:
			pro = pros[0]
			title = pro.find("p", attrs={"class":"title"})
			if getNodeText(title).find(cas) > -1:
				pLink = title.find("a")
				getProductInfo(pLink["href"], cas, products)
			else:
				products.append({"cas":cas})
		else:
			products.append({"cas":cas})
	else:
		products.append({"cas":cas})




def theardFun(name, inx):
	products = []
	excelFileName="ncbi"+str(inx)+".xlsx"
	wb = Workbook()
	workSheet = wb.active
	fileName="cat"+str(inx)+".txt"
	with open(fileName,'r') as file_to_read:
		index = 1
		type=1
		while True:
			lines = file_to_read.readline()
			if not lines:
					break
			getProductList(lines.replace("\n",""), products)

	headers=['cas','Product Name','Molecular Formula','Synonyms','Molecular Weight','IUPACName','InChI', 'InChIKey', 'CanonicalSMILES','CAS1','MolecularFormula']
	rindex = 1
	for p in products:
		writeExcel(workSheet, headers, rindex, p)
		rindex = rindex + 1
	print("flish")

	wb.save(excelFileName)
	
try:
	_thread.start_new_thread( theardFun, ('', 1 ) )
	_thread.start_new_thread( theardFun, ('', 2 ) )
	_thread.start_new_thread( theardFun, ('', 3 ) )
except:
	print ("Error: 无法启动线程")
while 1:
   pass